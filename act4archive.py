import os
import sys
import glob
import time
import datetime
from datetime import timedelta
import hashlib
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl import load_workbook
import docx

""" $env:MY_SOURCE_PATH
    $env:MY_RESULT_PATH
    $env:MY_SKR_VALIDATOR
    $env:MY_SET_FILE_PROPERTY = "YES"
"""

# types of files to be processed, skip the others
FILE_EXTENSION = {'.xlsx', '.xlsm', '.xls', '.docx', '.doc'}

# types of files to be processed, skip the others
EXEL_FILE_EXTENSION = {'.xlsx', '.xlsm', '.xls'}

""" you need to skip this folder:
00 Письма-запросы
00.1 Письма ГБ
01 ОСВ и первичка
02
10 АЗ и бух. отчетность
"""
NAME_SKIP_FOLDER = {'V:\\00 ', 'V:\\00.', 'V:\\01 ', 'V:\\02 ', 'V:\\10 '}

""" Это список папок, кот.нужно обрабатывать:
03 Инвентаризация  >>> no information about the author
04 Общая стратегия аудита
05 План проводимого аудита 
06 Аудит по существу
07 Предпосылки
08 Бухотчетность
09 Отчет аудитора
"""
# Строка поиска = "Проверил,\n выполнил " закодированная в utf-8, потому что есть перевод строки
KEYWORD1 = b'\xd0\x9f\xd1\x80\xd0\xbe\xd0\xb2\xd0\xb5\xd1\x80\xd0\xb8\xd0\xbb,\n \xd0\xb2\xd1\x8b\xd0\xbf\xd0\xbe\xd0\xbb\xd0\xbd\xd0\xb8\xd0\xbb '

# name of output file
OUT_FILE_NAME = r'Акт сдачи в архив электронных документов.xlsx'

# Руководитель Контрольной Службы
SKR_VALIDATOR_CONST = 'Атабиева М.И.'


def set_rightly_file_property(file, project_name, author, validator):

    boss = validator
    # check the current list item "file_list" this is a file?
    if not os.path.isfile(file):
        print('"This is not file": {0}'.format(file))
        # this is not a file - we proceed to the next iteration of the loop
        return -20    
    print('file= ', file)
    
    find_dot = file.rfind('.')
    if find_dot == -1:
        print('i do not find dot.')
        return -21

    ext_file = file[find_dot:]
    if ext_file == '.xlsx':
        print("======== PRINT func set_rightly_file_property: ============ \n file=", file, "\n project_name=", project_name, "\n author=", author, "\n boss=", boss)
        set_xlsx_file_property(file, project_name, author, boss)
    elif ext_file == '.docx':
        set_docx_file_property(file, project_name, author, boss)
            

def set_xlsx_file_property(file_name, project_name, author, boss):
    wb = load_workbook(file_name)
    # блок Описание
    wb.properties.title = "ООО Группа Финансы" # Название
    wb.properties.subject = project_name # Тема, можно поставить имя папки проекта
    wb.properties.keywords = "Нет" # Теги, можно поставить "Нет"
    wb.properties.category ="Нет" # Категории, можно поставить "Нет"
    wb.properties.description = "Нет" # Комментарии, можно поставить "Нет"
    # блок Источник
    wb.properties.creator = author #Авторы, можно поставить "ФИО руководителя проверки"
    wb.properties.lastModifiedBy = boss # Кем сохранён, можно поставить "ФИО руководителя проверки"
    wb.properties.revision = "001" # Редакция, можно поставить "Нет"
    wb.properties.version = "002" # Номер версии, можно поставить "Нет"
    # блок Источник
    wb.properties.created=datetime.datetime(2019, 1, 31, 9, 0, 0) # Дата создания содержимого
    wb.properties.modified=datetime.datetime(2019, 2, 28, 9, 0, 0) # Дата последнего сохранения
    wb.properties.lastPrinted=datetime.datetime(2019, 3, 30, 9, 0, 0) # Последний вывод на печать
    # блок Содержание
    wb.properties.contentStatus="None" # Состояние содержимого
    wb.properties.language="RUS" # Язык
    # Визуально Никуда не идёт: wb.properties.identifier = "None"
    #debug-print(wb.properties)
    
    wb.save(file_name)
    #time.sleep(1)


def set_docx_file_property(file_name, project_name, author, boss):
    doc = docx.Document(file_name)
    # блок Описание
    doc.core_properties.title = "ООО Группа Финансы" # Название, можно поставить имя папки "Авт сдачи в архив"
    doc.core_properties.subject = project_name # Тема, можно поставить имя папки проекта
    doc.core_properties.keywords = "Нет" # Теги, можно поставить "Нет"
    doc.core_properties.category = "Нет" # Категории, можно поставить "Нет"
    doc.core_properties.comments = "Нет" # Комментарии, можно поставить "Нет"
    # блок Источник
    doc.core_properties.author = author #Авторы, можно поставить "ФИО руководителя проверки"
    doc.core_properties.last_modified_by = boss # Кем сохранён, можно поставить "ФИО руководителя проверки"
    doc.core_properties.revision = 1 # Редакция, можно поставить "Нет"
    doc.core_properties.version= "002" # Номер версии, можно поставить "Нет"
    # блок Источник
    doc.core_properties.created = datetime.datetime(2020, 1, 31, 9, 0, 0) # Дата создания содержимого
    doc.core_properties.modified = datetime.datetime(2020, 2, 28, 9, 0, 0) # Дата последнего сохранения
    doc.core_properties.last_printed = datetime.datetime(2020, 3, 30, 9, 0, 0) # Последний вывод на печать
    # блок Содержание
    doc.core_properties.content_status ="None" # Состояние содержимого
    doc.core_properties.language='RUS' # Язык
    # Визуально Никуда не идёт: doc.core_properties.identifier='???'
    #debug-print(doc.core_properties)
    doc.save(file_name)
    time.sleep(1)


def get_auditors_list(file):
    """ returns the auditors(author, validator) who worked on the audit
        if not found then return 'НЕ найден'
    """

    NOT_FOUND = " --- "
    # checking the file's existence
    find_dot = file.rfind('.')
    if not (file[find_dot:] in EXEL_FILE_EXTENSION):
        # Не xls-расширение у файла
        return NOT_FOUND, NOT_FOUND

    #debug-print('source_book= ', file)
    # checking the existence of the sheet in the file
    wb = openpyxl.load_workbook(file)
    if wb.sheetnames.count('00') == 0:
        # Не найден лист '00'
        return NOT_FOUND, NOT_FOUND
    
    # читаем из excel-файла ФИО аудиторов
    # делаем лист '00' активным; row = строка, , column = столбец
    sheet = wb['00']
    found_author = False
    for row_count in range(17, 18):
        cell_value = sheet.cell(row=row_count, column=2).value
        if not(isinstance(cell_value, str)):
            continue
        template = cell_value.encode("utf-8")
        if template == KEYWORD1:
            cell_value = sheet.cell(row=row_count, column=3).value
            if isinstance(cell_value, str):
                validator = cell_value
            else:
                validator = NOT_FOUND
            author = validator
            if sheet.cell(row=row_count+1, column=2).value == "Заполнил":
                cell_value =  sheet.cell(row=row_count+1, column=3).value
                if isinstance(cell_value, str):
                    author = cell_value

            found_author = True
            break
    
    if found_author:
        return author, validator
    else:
        return NOT_FOUND, NOT_FOUND           


def get_source_path():
    """ returns the path to the folder where the data is stored (files)
    """
    if "MY_SOURCE_PATH" in os.environ:
        # имя каталога должно быть без кавычек
        return os.environ["MY_SOURCE_PATH"].replace('"', '')
    else:
        return os.getcwd()


def get_result_path():
    """ returns the path to the folder where you want to save data (files)
    """
    if "MY_RESULT_PATH" in os.environ:
        # имя каталога должно быть без кавычек
        return os.environ["MY_RESULT_PATH"].replace('"', '')
    else:
        # save to the current folder
        return os.getcwd()    

def get_skr_validator():
    """ returns 'Руководитель Контрольной Службы'
    """
    if "MY_SKR_VALIDATOR" in os.environ:
        # имя должно быть без кавычек
        return os.environ["MY_SKR_VALIDATOR"].replace('"', '')
    else:
        return SKR_VALIDATOR_CONST
    

def file_is_needed(file):
    """ Check - this file needs to be processed. If need, then return True,
    otherwise - False.
    """
    find_dot = file.rfind('.')
    if find_dot == -1:
        return False
    
    elif (file[find_dot:] in FILE_EXTENSION):
        find_temp = file.rfind('\\~$')
        if find_temp == -1:
            return True
        else:
            # This is temporary file, skip'
            return False
    else:
        return False

def skip_this_folder(file):
    """ Check - you need to skip this folder. If need, then return True,
    otherwise - False.
    """
    return (file[0:6] in NAME_SKIP_FOLDER)


def remove_link_to_v():
    """ Remove the Symbolic link to disk V: """
    if os.path.exists("V:\\"):
        subst_command = 'subst /d V:'
        return os.system(subst_command)


def get_hash_md5(filename):
    """ get hash-MD5 of filename object """

    with open(filename, 'rb') as f:
        m = hashlib.md5()
        while True:
            data = f.read(8192)
            if not data:
                break
            m.update(data)
        return m.hexdigest()


def get_dirty_file_list(source_path):
    """ get a list of all items in the current directory
    """

    # Удаляем Символическую ссылку на диск V: если она существует
    remove_link_to_v()
    # Подключаем Симвлолическую ссылку диск V: указывает на "source_path"
    subst_command = 'subst V: "{0}"'.format(source_path)

    ret_code = os.system(subst_command)
    if ret_code != 0:
        print('Символ.ссылка НЕ СДЕЛАНА! V:-> {0} Код возврата subst: {1}'.
              format(source_path, ret_code))
        return None
    else:
        print('Символ.ссылка СДЕЛАНА диск V: -> {0} Код возврата subst: {1}'.
              format(source_path, ret_code))

    # создаём рекурсивно список объектов внутри "V:\ (src_dir_level1)"
    return glob.glob('V:\\**\\*', recursive=True)


def create_table_and_set_file_property(
        file_list,
        project_name):
    """ create а table from a 'file_list' each row consists of:
    file name + file hash + author + validator
    Если установлена переменная среды os.environ["MY_SET_FILE_PROPERTY"] == "YES",
    то устанавливаем у файлов doc и xls нужные свойства файла
    """

    if ("MY_SET_FILE_PROPERTY" in os.environ) and (os.environ["MY_SET_FILE_PROPERTY"] == "YES"):
        set_file_prop = True
    else:
        set_file_prop = False

    result_table = []
    for file in file_list:
        # check the current list item "file_list" this is a file?
        if not os.path.isfile(file):
            print('We need to skip because it is folder and not file: {0}'.format(file))
            # this is not a file - we proceed to the next iteration of the loop
            continue
        if not file_is_needed(file):
            print('We need to skip because this file is not needed: {0}'.format(file))
            # we skip this file and proceed to the next iteration of the loop
            continue
        if skip_this_folder(file):
            print('We need to skip files in this folder: {0}'.format(file))
            # we skip this file and proceed to the next iteration of the loop
            continue
        print('обрабатываем файл: {0}'.format(file))

        author, validator = get_auditors_list(file)
        if set_file_prop:
            set_rightly_file_property(file, project_name, author, validator)
        curr_line = {}
        curr_line.update({'file-name': file[2:]})
        # modtime = time.localtime(os.path.getmtime(file))
        # curr_line['file_modify'] =time.strftime("%d-%m-%Y %H:%M:%S", modtime)
        curr_line.update({'Hash-MD5': get_hash_md5(file)})
        curr_line['author'] = author          # Автор==Заполнил
        curr_line['validator'] = validator    # Проверил
        result_table.append(curr_line)
    return result_table


def save_table_to_xlsx_file(project_name, result_table, result_path, skr_validator):
    """ write 'result_table' to xlsx file """

    wb = openpyxl.Workbook()
    wb.create_sheet(title='Список файлов', index=0)
    sheet = wb['Список файлов']
    # create a header
    col = 1
    row = 1
    sheet.merge_cells('A1:E1')
    sheet['A1'] = 'АКТ'
    sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
    sheet['A1'].font = Font(bold=True)
    sheet.merge_cells('A2:E2')
    sheet['A2'] = 'сдачи в архив электронных документов по проекту ' + project_name
    sheet['A2'].alignment = Alignment(horizontal="center", vertical="center")
    sheet['A2'].font = Font(bold=True)
    # create a table-header row
    bbb = Side(border_style="medium", color="000000")
    sheet['A4'] = '№ п/п'
    sheet.column_dimensions['A'].width = 7
    sheet['A4'].font = Font(bold=True)
    sheet['A4'].border = Border(top=bbb, left=bbb, right=bbb, bottom=bbb)

    sheet['B4'] = 'Имя файла'
    sheet.column_dimensions['B'].width = 100
    sheet['B4'].font = Font(bold=True)
    sheet['B4'].border = Border(top=bbb, left=bbb, right=bbb, bottom=bbb)
    # sheet['C1'] = 'дата последнего изменения'
    sheet['C4'] = 'Хэш-сумма (MD5) файла'
    sheet.column_dimensions['C'].width = 37
    sheet['C4'].font = Font(bold=True)
    sheet['C4'].border = Border(top=bbb, left=bbb, right=bbb, bottom=bbb)

    sheet['D4'] = 'Файл заполнил'
    sheet.column_dimensions['D'].width = 20
    sheet['D4'].font = Font(bold=True)
    sheet['D4'].border = Border(top=bbb, left=bbb, right=bbb, bottom=bbb)

    sheet['E4'] = 'Файл проверил'
    sheet.column_dimensions['E'].width = 20
    sheet['E4'].font = Font(bold=True)
    sheet['E4'].border = Border(top=bbb, left=bbb, right=bbb, bottom=bbb)

    row = 4
    thin = Side(border_style="thin", color="000000")
    for line_dict in result_table:
        row += 1
        col = 1
        cell = sheet.cell(row=row, column=col)
        # это номер строки таблицы с файлами
        cell.value = row-4
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for __, val in line_dict.items():
            col += 1
            cell = sheet.cell(row=row, column=col)
            cell.value = val
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    col = 1
    cell = sheet.cell(row=row+4, column=col)
    cell.value = f'Файлы принял на хранение в архив'
    cell = sheet.cell(row=row+5, column=col)
    cell.value = f'Руководитель Контрольной Службы:'
    cell = sheet.cell(row=row+5, column=col+3)
    cell.value = f'{skr_validator}'
    
    file_name = result_path + '\\' + project_name + '_' + OUT_FILE_NAME
    wb.save(file_name)


def process_create_file_table(source_path, result_path, skr_validator):
    """ Creates a list of all files as: file name, hash, author """

    file_list = get_dirty_file_list(source_path)
    if file_list is None:
        print('A list of all items in current directory is not created, RC=-10')
        return -10
  
    project_name = source_path[source_path.rindex('\\') + 1:]

    result_table = create_table_and_set_file_property(
        file_list,
        project_name
        )
    
    save_table_to_xlsx_file(
        project_name,
        result_table,
        result_path,
        skr_validator
        )
  
    # Удаляем Симвлолическую ссылку на диск V:
    remove_link_to_v()
    return 0

def main(argv=None):
    start_time = time.monotonic()
    source_path = get_source_path()
    print("source_path = ", source_path)

    result_path = get_result_path()
    print ("result_path = ", result_path)

    skr_validator = get_skr_validator()
    print("skr_validator = ", skr_validator)
    
    ret_code = process_create_file_table(
        source_path,
        result_path,
        skr_validator
    )
    end_time = time.monotonic()
    print("Сгенерирован акт в каталоге ", result_path,
         "\nпо проекту ", source_path, 
         "\nВремя выполнения: ", timedelta(seconds=end_time - start_time))
    print("return code =", ret_code)
    return ret_code


if __name__ == '__main__':
    sys.exit(main(sys.argv))
