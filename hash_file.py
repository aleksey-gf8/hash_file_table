import os
# import time
import glob
import hashlib
from openpyxl.styles import Border, Side, Font, Alignment
import openpyxl


# types of files to be processed, skip the others
file_extension = {'.xlsx', '.xls', '.docx', '.doc'}

# skip_this_folder
name_skip_folder = {'V:\\00 ', 'V:\\01 '}

# global var
auditors_list = []

OUT_FILE_NAME = r'Акт сдачи в архив электронных документов'
SKR_VALIDATOR = 'Усков Д.И.'


def get_auditors_list(source_path):
    """ returns a list of auditors who worked on the audit
        taken from a file ".\05  Аудит  СВК\05.00 Содержание.xlsx"
    """
    source_books = ["V:\\05  Аудит  СВК\\05.00 Содержание.xlsx",
                    "V:\\05 Аудит СВК\\05.00 Содержание.xlsx",
                    "V:\\05 Аудит  СВК\\05.00 Содержание.xlsx",
                    "V:\\05  Аудит СВК\\05.00 Содержание.xlsx",
                    ]
    not_found = True
    for source_book in source_books:
        if not os.path.isfile(source_book):
            continue
        else:
            not_found = False
            break

    if not_found:
        print('"This file does not exist": {0}'.format(source_book))
        return None
    # читаем из excel-файла ФИО аудиторов
    wb = openpyxl.load_workbook(source_book)
    if wb.sheetnames.count('12') == 0:
        print('"Sheet name 12 does not exist": {0}'.format(source_book))
        return None
    result = []
    # делаем 12 лист активным
    wb.active = 11
    # получаем активный лист
    sheet = wb.active
    if sheet['A29'].value == 1:
        result.append(sheet['B29'].value)
    if sheet['A31'].value == 'Руководитель проверки:':
        result.append(sheet['B31'].value)
    else:
        numbers_by_row = [32, 33, 34, 35]
        for row_count in numbers_by_row:
            if sheet.cell(row=row_count, column=1).value == 'Руководитель проверки:':
                result.append(sheet.cell(row=row_count, column=2).value)
                break
    print('result= ', result)
    return result


def get_source_path():
    """ returns the path to the folder where the data is stored (files)
    """
    if "MY_SOURCE_PATH" in os.environ:
        # имя каталога должно быть без кавычек
        return os.environ["MY_SOURCE_PATH"].replace('"', '')
    else:
        return os.getcwd()


def file_is_needed(file):
    """ Check - this file needs to be processed. If need, then return True,
    otherwise - False.
    """
    find_dot = file.rfind('.')
    if find_dot == -1:
        return False
    return (file[find_dot:] in file_extension)


def skip_this_folder(file):
    """ Check - you need to skip this folder. If need, then return True,
    otherwise - False.
    """
    return (file[0:6] in name_skip_folder)


def remove_link_to_v():
    """ Remove the Symbolic link to disk V: """
    if os.path.exists("V:\\"):
        subst_command = 'subst /d V:'
        return os.system(subst_command)


def get_hash_md5(filename):
    """ get hash-MD5 of filename object """

    with open(filename, 'rb') as f:
        m = hashlib.md5()
        while True:
            data = f.read(8192)
            if not data:
                break
            m.update(data)
        return m.hexdigest()


def get_file_list(source_path):
    """ get a list of all items in the current directory
    """

    # Удаляем Символическую ссылку на диск V: если она существует
    remove_link_to_v()
    # Подключаем Симвлолическую ссылку диск V: указывает на "source_path"
    subst_command = 'subst V: "{0}"'.format(source_path)
    print(subst_command)
    ret_code = os.system(subst_command)

    if ret_code != 0:
        print('Символ.ссылка НЕ СДЕЛАНА! V:-> {0} Код возврата subst: {1}'.
              format(source_path, ret_code))
        # завершаем и переходим к следующей итерации цикла по source_dir.
        return None
    else:
        print('Символ.ссылка СДЕЛАНА диск V: -> {0} Код возврата subst: {1}'.
              format(source_path, ret_code))

    # создаём рекурсивно список объектов внутри "V:\ (src_dir_level1)"
    return glob.glob('V:\\**\\*', recursive=True)


def create_table(file_list):
    """ create а table from a 'file_list' each row consists of:
    file name + file hash + author + validator
    """
    global auditors_list
    auditors_list = get_auditors_list(source_path)
    if auditors_list:
        author = auditors_list[0]
        validator = auditors_list[1]
    else:
        print("no read author")
        author = 'author'
        validator = 'validator'
    result_table = []
    for file in file_list:
        # check the current list item "file_list" this is a file?
        if not os.path.isfile(file):
            print('"This is not file": {0}'.format(file))
            # this is not a file - we proceed to the next iteration of the loop
            continue
        if not file_is_needed(file):
            print('"This file is not needed": {0}'.format(file))
            # we skip this file and proceed to the next iteration of the loop
            continue
        if skip_this_folder(file):
            print('"you need to skip this folder": {0}'.format(file))
            # we skip this file and proceed to the next iteration of the loop
            continue
        print('обрабатываем файл: {0}'.format(file))
        curr_line = {}
        curr_line.update({'file-name': file[2:]})
        # modtime = time.localtime(os.path.getmtime(file))
        # curr_line['file_modify'] =time.strftime("%d-%m-%Y %H:%M:%S", modtime)
        curr_line.update({'Hash-MD5': get_hash_md5(file)})
        curr_line['author'] = author
        curr_line['validator'] = validator
        result_table.append(curr_line)
    return result_table


def save_table_to_txt_file(result_table, result_path):
    """ write 'result_table' to text file """

    f_name = result_path + '.txt'
    with open(f_name, "w", encoding='utf-8') as file:
        for line_dict in result_table:
            for key, val in line_dict.items():
                file.write('{}: {}\t'.format(key, val))
            file.write('\n')
        file.write(f'\n\t Исходный каталог = \"{source_path}\"\n')
        file.write(f'\n\t Количество файлов = \"{len(result_table)}\"\n')


def save_table_to_xlsx_file(result_table, result_path):
    """ write 'result_table' to xlsx file """

    wb = openpyxl.Workbook()
    wb.create_sheet(title='Список файлов', index=0)
    sheet = wb['Список файлов']
    # create a header
    col = 1
    row = 1
    sheet.merge_cells('A1:E1')
    sheet['A1'] = 'АКТ'
    sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
    sheet['A1'].font = Font(bold=True)
    sheet.merge_cells('A2:E2')
    sheet['A2'] = 'сдачи в архив электронных документов'
    sheet['A2'].alignment = Alignment(horizontal="center", vertical="center")
    sheet['A2'].font = Font(bold=True)
    # create a table-header row
    bbb = Side(border_style="medium", color="000000")
    sheet['A4'] = '№ п/п'
    sheet.column_dimensions['A'].width = 7
    sheet['A4'].font = Font(bold=True)
    sheet['A4'].border = Border(top=bbb, left=bbb, right=bbb, bottom=bbb)

    sheet['B4'] = 'Имя файла'
    sheet.column_dimensions['B'].width = 100
    sheet['B4'].font = Font(bold=True)
    sheet['B4'].border = Border(top=bbb, left=bbb, right=bbb, bottom=bbb)
    # sheet['C1'] = 'дата последнего изменения'
    sheet['C4'] = 'Хэш-сумма (MD5) файла'
    sheet.column_dimensions['C'].width = 37
    sheet['C4'].font = Font(bold=True)
    sheet['C4'].border = Border(top=bbb, left=bbb, right=bbb, bottom=bbb)

    sheet['D4'] = 'Файл создал'
    sheet.column_dimensions['D'].width = 20
    sheet['D4'].font = Font(bold=True)
    sheet['D4'].border = Border(top=bbb, left=bbb, right=bbb, bottom=bbb)

    sheet['E4'] = 'Файл проверил'
    sheet.column_dimensions['E'].width = 20
    sheet['E4'].font = Font(bold=True)
    sheet['E4'].border = Border(top=bbb, left=bbb, right=bbb, bottom=bbb)

    row = 5
    thin = Side(border_style="thin", color="000000")
    for line_dict in result_table:
        row += 1
        col = 1
        cell = sheet.cell(row=row, column=col)
        # this is a row counter, you need to take into account the empty rows
        # in the table header
        cell.value = row-5
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for __, val in line_dict.items():
            col += 1
            cell = sheet.cell(row=row, column=col)
            cell.value = val
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    col = 1
    # validator = auditors_list[1] if auditors_list else 'validator'
    cell = sheet.cell(row=row+4, column=col)
    cell.value = f'Файлы принял на хранение в архив'
    cell = sheet.cell(row=row+5, column=col)
    cell.value = f'Руководитель Контрольной Службы:'
    cell = sheet.cell(row=row+5, column=col+3)
    cell.value = f'{SKR_VALIDATOR}'

    f_name = result_path + '.xlsx'
    wb.save(f_name)


def process_create_file_table(source_path, result_path):
    """ Creates a list of all files as: file name, hash, author """

    file_list = get_file_list(source_path)
    if file_list is None:
        print('A list of all items in current directory is not created')
        return None
    result_table = create_table(file_list)

    # save_table_to_txt_file(result_table, result_path)
    save_table_to_xlsx_file(result_table, result_path)

    # Удаляем Симвлолическую ссылку на диск V:
    remove_link_to_v()


if __name__ == '__main__':
    source_path = get_source_path()
    print(source_path)
    result_path = OUT_FILE_NAME
    process_create_file_table(
        source_path,
        result_path
    )
